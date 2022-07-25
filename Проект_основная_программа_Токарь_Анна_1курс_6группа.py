import telebot
import openpyxl
import os
import random
from telebot import types

bot = telebot.TeleBot(token='1676769888:AAEZVSV2A4KgO85p4RO4l7u3wyxwUWCR4K8')
summa_zakaza = 0  # общая сумма заказа
nomer_tovara = 0  # ключи в корзине (словаре)
corzina = {}        # словарь с товарами
corzina_polz = {}   # словарь с ключём - chat.id и corzina{}

# пояснение работы программы после каждой функции
@bot.message_handler(commands=['start'])
def first(message):
    keyboard = types.ReplyKeyboardMarkup()
    keyboard.add('Меню🍺🌭')
    keyboard.add('Акции')
    keyboard.add('Заказ🛍')
    keyboard.add('Связь с нами📞')
    photo = open('foto/' + random.choice(os.listdir('foto')), 'rb')
    bot.send_photo(message.from_user.id, photo)
    send = bot.send_message(message.chat.id, 'Привет!🤗 Я - бот-доставка или просто НеотЛожка! \n Просто выберите, что бы Вы хотели заказать.',reply_markup=keyboard)
    bot.register_next_step_handler(send, second)
    '''
    @bot.message_handler(commands=['start']) - ожидает сообщения-команды (которые начинаются с /). Здесь ожидает команду /start
    keyboard - массив кнопок (объекты KeyboardButton)
    ReplyKeyboardMarkup() — объект, который создает клавиатуру
    keyboard.add - добавление кнопки
    listdir() модуля os возвращает список, содержащий имена файлов и директорий в каталоге, заданном путем foto.
    send_photo и send_message - отправка фото и сообщения пользователю
    register_next_step_handler(send, second) - принимает два аргумента: первый - message, второй - function; он ждёт сообщение пользователя и потом вызывает указанную функцию с аргументом message
    '''

def second(message):
    keyboard = types.ReplyKeyboardMarkup()
    markup_inline = types.InlineKeyboardMarkup()
    if message.text == 'Меню🍺🌭':
        keyboard.row('Что выпить?🍻','Что поесть?🍗')
        keyboard.row('Фирменная продукция', 'Всё для шашлыка')
        keyboard.add('Заказ🛍')
        keyboard.add('Назад🔚')
        send = bot.send_message(message.chat.id,'Выберете:', reply_markup=keyboard)
        bot.register_next_step_handler(send,third)
    elif message.text == 'Акции':
        four(message)
    elif message.text == 'Связь с нами📞':
        inst_button = types.InlineKeyboardButton(text="Инстаграм @pinta_market", url="https://www.instagram.com/pinta_market/")
        poisk_button = types.InlineKeyboardButton(text="Адреса Pinta Market", url="https://goo.su/4fGs")
        markup_inline.add(inst_button, poisk_button)
        send = bot.send_message(message.chat.id, "Привет! Нажми на кнопку и перейди на наш инстаграм или карту с нашими магазинами.",reply_markup=markup_inline)
        keyboard.add('Назад🔚')
        first(message)
    elif message.text == 'Заказ🛍':
        seven(message, corzina)
    '''
    message.text - текст отправляемого сообщения, 1-4096 знаков
    .row — метод для группировки кнопок в строку.
    keyboard.row и keyboard.add - добавление кнопки
    InlineKeyboard — клавиатура привязанная к сообщению, изпользующая обратный вызов (CallbackQuery), вместо отправки сообщения с обыкновенной клавиатуры.
    url - это адрес, указывающий путь к интернет ресурсу
    '''

def third(message):
    keyboard = types.ReplyKeyboardMarkup()
    if message.text == 'Что выпить?🍻':
        keyboard.row('Пиво - бариэль', 'Пиво - крафт')
        keyboard.row('Кроп - пиво', 'Пиво - лидское')
        keyboard.row('Пиво - крюгер', 'Пиво - Балтика')
        keyboard.row('Пиво - партнер', 'Пиво - свет.город')
        keyboard.row('Пиво - пинта', 'Пиво - другое')
        keyboard.row('Безалкогольные напитки')
        keyboard.add('Назад🔚')
        send = bot.send_message(message.chat.id,'🍻 Выбор остаётся за Вами:', reply_markup=keyboard)
        bot.register_next_step_handler(send,four)
    elif message.text == 'Что поесть?🍗':
        keyboard = types.ReplyKeyboardMarkup(True,False)
        keyboard.row('Детские сладости', "Чипсы LAY'S и Бруто")
        keyboard.row('Мясо и сыр', 'Орехи')
        keyboard.row('Рыбка', 'Кальмары')
        keyboard.row('Семечки', 'Сухари и гренки', 'Пинта - снеки')
        keyboard.add('Назад🔚')
        send = bot.send_message(message.chat.id,'🍗 Выберите чем бы подкрепиться:', reply_markup=keyboard)
        bot.register_next_step_handler(send,four)
    elif message.text == 'Фирменная продукция' or message.text == 'Всё для шашлыка':
        four(message)
    elif message.text == 'Назад🔚':
        first(message)
    elif message.text == 'Заказ🛍':
        seven(message, corzina)
    else:
        bot.send_message(message.chat.id,'Извините, такой функции нет, вернитесь к началу.')
        first(message)
    '''
    four(message) - открытие функции four после выбора кнопкок
     'Фирменная продукция' или 'Всё для шашлыка'
    first(message) - возвращение пользователя в основное меню после выбора кнопки 'Назад🔚'
    seven(message, corzina) - открытие функции seven после выбора кнопки 'Заказ🛍'
    '''

def four(message):
    if message.text == 'Назад🔚':
        message.text = 'Меню🍺🌭'
        second(message)
    else:
        markup = telebot.types.InlineKeyboardMarkup()
        keyboard = types.ReplyKeyboardMarkup()
        t = message.text + '.xlsx'
        wb = openpyxl.load_workbook(t)
        sheet = wb.active
        rows = sheet.max_row
        cols = sheet.max_column
        for i in range(1, rows + 1):
            string = ''
            for j in range(1, cols):
                cell = sheet.cell(row = i, column = j)
                string = string + str(cell.value)  + '  -> цена:\n   '
            keyboard.add(string[:-12])
        keyboard.add('Назад🔚')
        photo = open('foto/' + random.choice(os.listdir('foto')), 'rb')
        buy = bot.send_photo(message.from_user.id, photo)
        bot.send_message(message.chat.id, message.text + ':', reply_markup=keyboard)
        bot.register_next_step_handler(buy, five_tovar_gr, t, corzina_polz)
    '''
    openpyxl — это библиотека Python для чтения и записи файлов Excel
    load_workbook() принимает имя файла в качестве аргумента и возвращает объект рабочей книги, который представляет файл.
    wb = openpyxl.load_workbook(t) - открываем тестовый Excel файл
    sheet = wb.active - получаем активный лист
    max_row и max_column - размер листа
    t - название выбраннного файла excel
    '''

def five_tovar_gr(message, t, corzina_polz):
    keyboard = types.ReplyKeyboardMarkup()
    global nomer_tovara
    if message.text == 'Назад🔚':
        message.text = 'Меню🍺🌭'
        second(message)
    elif message.text == 'Заказ🛍':
        corzina.append(message.text)
        eight(message)
    else:
        nomer_tovara += 1
        corzina[nomer_tovara] = message.text
        corzina_polz = {message.chat.id: corzina}
        print(corzina_polz)
        photo = open('foto/' + random.choice(os.listdir('foto')), 'rb')
        bot.send_photo(message.from_user.id, photo)
        a = telebot.types.ReplyKeyboardRemove()
        if (t == 'Мясо и сыр.xlsx' or t == 'Орехи.xlsx' or t == 'Рыбка.xlsx' or t == 'Кальмары.xlsx'
        or t =='Сухари и гренки.xlsx' or t == 'Пинта - снеки.xlsx'):
            bot.send_message(message.from_user.id, '⚖️ Сколько граммов? \n🌐 Напишите цифрами, пожалуйста', reply_markup=a)
        else:
            bot.send_message(message.from_user.id, '📦 Сколько штук? \n🌐 Напишите цифрами, пожалуйста', reply_markup=a)
        bot.register_next_step_handler(message, five, t, corzina_polz)
    '''
    telebot.types.ReplyKeyboardRemove() - для удаления клавиатуры
    (чтобы пользователь сам выбрал количество товара)
    corzina[nomer_tovara] = message.text - добавление товара в корзину
    corzina_polz = {message.chat.id: corzina} - словарь с ключём - message.chat.id и корзиной
    '''

def five(message, t, corzina_polz):
    keyboard = types.ReplyKeyboardMarkup()
    kol_tovara = 0
    if message.text.isdigit() == False:
        bot.send_message(message.from_user.id,'🌐 Напишите цифрами, пожалуйста', reply_markup=keyboard)
        message.text = corzina_polz[message.chat.id][nomer_tovara]
        del corzina_polz[message.chat.id][nomer_tovara]
        print(message.text)
        print(corzina_polz)
        five_tovar_gr(message, t, corzina_polz)
    else:
        kol_tovara = int(message.text)
        if (t == 'Мясо и сыр.xlsx' or t == 'Орехи.xlsx' or t == 'Рыбка.xlsx' or t == 'Кальмары.xlsx'
            or t =='Сухари и гренки.xlsx' or t == 'Пинта - снеки.xlsx'):
            question = 'Вам '+ str(kol_tovara) +' граммов?'
            kol_tovara = kol_tovara / 1000
            keyboard.row('Да, далее✅', 'Нет, другое❌')
        else:
            question = 'Вам '+ str(kol_tovara) +' штук/штуки?'
            keyboard.row('Да✅', 'Нет❌')
        keyboard.row('Убрать из корзины❌')
        bot.send_message(message.from_user.id, question, reply_markup=keyboard)
        bot.register_next_step_handler(message, six, corzina_polz, kol_tovara)
    '''
    kol_tovara - количество товара
    kol_tovara = kol_tovara / 1000 - для счёта в граммах
    '''

def six(message, corzina_polz, kol_tovara):
    keyboard = types.ReplyKeyboardMarkup()
    global nomer_tovara
    if message.text == 'Нет, другое❌' or message.text == 'Нет❌':
        message.text = corzina_polz[message.chat.id][nomer_tovara]
        corzina_polz[message.chat.id].popitem()
        t = message.text
        five_tovar_gr(message, t, corzina_polz)
    elif message.text == 'Убрать из корзины❌':
        bot.send_message(message.chat.id, '📌 Вы удалили из корзины: \n' + str(corzina_polz[message.chat.id].popitem()), reply_markup=keyboard)
        keyboard.add('Заказать еще🍕')
        keyboard.add('Заказ🛍')
        keyboard.add('Отменить заказ❌')
        sent = bot.send_message(message.chat.id, 'Выберете функцию:', reply_markup=keyboard)
        bot.register_next_step_handler(sent, seven, corzina_polz)
    else:
        photo = open('foto/' + random.choice(os.listdir('foto')), 'rb')
        buy = bot.send_photo(message.from_user.id, photo)
        message.text = corzina_polz[message.chat.id][nomer_tovara][:-13]
        print(int(corzina_polz[message.chat.id][nomer_tovara][-4::1]))
        summa_tovara = ('%6.2f' %(float(kol_tovara) * int(corzina_polz[message.chat.id][nomer_tovara][-4::1])))
        k = (corzina_polz[message.chat.id][nomer_tovara] +  ' рублей' + ' * ' + str(kol_tovara) + ' =  ' + str(summa_tovara))
        corzina_polz[message.chat.id][nomer_tovara] = k
        print(corzina_polz)
        zak = '🛒 Вы добавили в корзину: \n' + k
        bot.send_message(message.chat.id, zak, reply_markup=keyboard)
        keyboard.add('Заказать еще🍕')
        keyboard.add('Заказ🛍')
        keyboard.add('Отменить заказ❌')
        sent = bot.send_message(message.chat.id, 'Выберете функцию:', reply_markup=keyboard)
        bot.register_next_step_handler(sent, seven, corzina_polz)
    '''
    summa_tovara - цена самого товара * на его количество (выбор пользователя)
    '''

def seven(message, corzina_polz):
    keyboard = types.ReplyKeyboardMarkup()
    global summa_zakaza
    if message.text == 'Заказать еще🍕':
        message.text = 'Меню🍺🌭'
        second(message)
    elif message.text == 'Заказ🛍':
        bot.send_message(message.chat.id, 'Ваша корзина🛒:', reply_markup=keyboard)
        print(corzina_polz)
        if corzina_polz == {}:
            bot.send_message(message.chat.id, 'Ваша корзина пустая...Вернитесь в меню.', reply_markup=keyboard)
            first(message)
        else:
            corzina_admin = 'Пользователь:  ' + str(message.chat.id) + '\n'
            for i in corzina_polz[message.chat.id].values():
                corzina_admin = corzina_admin + i + '\n'
                print(i[-6::1])
                summa_zakaza = summa_zakaza + float(i[-8::1])
            bot.send_message(message.chat.id, '💰 Общая сумма заказа:  ' + str(summa_zakaza) + ' рублей', reply_markup=keyboard)
            zakaz = bot.send_message(message.chat.id, corzina_admin, reply_markup=keyboard)
            bot.forward_message("@bot_pinta_market_rnd", message.chat.id, zakaz.message_id)
            nomer(message)
    elif message.text == 'Отменить заказ❌':
        bot.send_message(message.chat.id, '📌 Вы отменили свой заказ', reply_markup=keyboard)
        corzina_polz.clear()
        x = 0
        summa_zakaza = 0
        first(message)
    '''
    bot.forward_message("@bot_pinta_market_rnd", message.chat.id, zakaz.message_id) -
    отправка сообщения (товар из корзины) в другой чат (чат с администратором)
    summa_zakaza - общая сумма заказа
    '''

def nomer(message):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    button_phone = types.KeyboardButton(text="📌 Отправить номер телефона", request_contact=True)
    keyboard.add(button_phone)
    sent =(bot.send_message(message.chat.id, '✅ Отправьте свой номер телефона: \n Вы можете просто написать его', reply_markup=keyboard))
    bot.register_next_step_handler(sent, send_telegram_nom)
    '''
    resize_keyboard - если предать true, то клавиатура подгонится по высоте до возможного минимума.
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True) - кнопка с ее закрытием
    text - текст который будет на отображен на кнопке, обязательный параметр, поддерживает текст и смайлики (эмодзи)
    request_contact - если параметр установлен в true, пользователь отправит в бот свой номер телефона на который зарегистрирован аккаунт
    '''

def send_telegram_nom(message):
    bot.forward_message("@bot_pinta_market_rnd", message.chat.id, message.message_id)
    geodan(message)

def geodan(message):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    button_geo = types.KeyboardButton(text="📌 Отправить местоположение", request_location=True)
    keyboard.row(button_geo)
    if message.location is not None:
        print(message.location)
        print("latitude: %s; longitude: %s" % (message.location.latitude, message.location.longitude))
    sent = bot.send_message(message.chat.id, '✅ Отправьте местоположение, в которое нужно доставить заказ: \n Вы можете просто написать его.', reply_markup=keyboard)
    bot.register_next_step_handler(sent, send_telegram_geo)
    '''
    resize_keyboard - если предать true, то клавиатура подгонится по высоте до возможного минимума.
    text - текст который будет на отображен на кнопке, обязательный параметр, поддерживает текст и смайлики (эмодзи)
    request_location - если параметр установлен в true, пользователь отправит в бот свое текущее местоположение
    '''

def send_telegram_geo(message):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    bot.forward_message("@bot_pinta_market_rnd", message.chat.id, message.message_id)
    sent = bot.send_message(message.from_user.id, '✅ Заказ принят! Наш администратор свяжется с Вами в ближайшее время!', reply_markup=keyboard)
    send_messages_1(message)

def send_messages_1(message):
    global nomer_tovara
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    keyboard.add('📌 Новый заказ!')
    sent = bot.send_message(message.from_user.id, '🛍 Новый заказ! Нажмите на кнопку!', reply_markup=keyboard)
    nomer_tovara = 0
    bot.register_next_step_handler(sent, first)

@bot.message_handler(content_types=["text", "sticker", "pinned_message", "photo", "audio"])
def send_messages_2(message): # Название функции не играет никакой роли, в принципе
    bot.send_message(message.chat.id,'Такой функции не существует. Мы вернули Вас в начальное меню.')
    first(message)

bot.infinity_polling()